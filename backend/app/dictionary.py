"""
専門用語・類似語辞書 (Glossary / Synonyms) 管理モジュール
仕様:
- Excel (.xlsx) および CSV (.csv) ファイルから用語・類似語・解説を読み込みインメモリ管理。
- 新2列フォーマット（第1列: 専門用語（カンマ区切り同義語含む）, 第2列: 意味・解説）および従来の3列フォーマットの両方に完全対応。
- SHA-256 ハッシュおよび mtime/size メタデータによる差分検知 & インメモリキャッシュ機構（I/O負荷ゼロ化 & 高速応答）。
- 人間が外部でファイルを編集した際のみ自動差分検知して再パース・更新。
- 列名（日本語「専門用語」「類似語」「意味・解説」/ 英語「Term」「Synonyms」「Description」等）の自動判定。
- 大文字小文字、全角半角、ハイフン有無（PJX ⇔ PJ-X ⇔ ｐｊｘ）の表記揺れ正規化。
- 自然言語テキスト（質問文・ノート本文）からの最長一致用語検知。
- 自然言語クエリのEmbedding用セマンティック補強文字列の生成。
- チャンキング時における同義語・解説メタデータ抽出機能の提供。
- Web UIからの辞書作成・編集・Excel (.xlsx) 保存機能の提供。
"""

import csv
import hashlib
import os
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union


def normalize_term_key(text: str) -> str:
    """
    表記揺れを吸収するための正規化キーを生成する。
    - 全角英数を半角に変換 (NFKC)
    - 小文字化
    - ハイフン・アンダースコア・スペースの除去
    """
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKC", text).lower()
    normalized = re.sub(r"[\s\-_ー−―]+", "", normalized)
    return normalized


def compute_file_sha256(file_path: str) -> str:
    """ファイルのSHA-256ハッシュ値を高速計算する"""
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(65536):
            sha256.update(chunk)
    return sha256.hexdigest()


@dataclass
class GlossaryEntry:
    """専門用語エントリ"""
    term: str
    synonyms: List[str] = field(default_factory=list)
    description: str = ""
    # 表記揺れキー群（termおよび全synonymsの正規化後キー）
    normalized_keys: Set[str] = field(default_factory=set)

    def all_variants(self) -> List[str]:
        """代表語およびすべての類似語を含む重複なしリスト"""
        res = [self.term]
        for s in self.synonyms:
            if s and s not in res:
                res.append(s)
        return res

    def to_dict(self) -> Dict[str, Any]:
        """フロントエンド/API連携用の辞書形式へ変換"""
        variants = self.all_variants()
        return {
            "term": self.term,
            "synonyms": self.synonyms,
            "description": self.description,
            "terms": ", ".join(variants),
        }


class GlossaryDictionary:
    """専門用語・類似語辞書クラス（ハッシュ差分キャッシュ機能付き）"""

    # ファイルパスごとのインメモリキャッシュ: {abs_path: {"mtime": float, "size": int, "sha256": str, "instance": GlossaryDictionary}}
    _cache: Dict[str, Dict[str, Any]] = {}

    def __init__(
        self,
        entries: Optional[List[GlossaryEntry]] = None,
        file_path: str = "",
        file_sha256: Optional[str] = None,
        file_mtime: Optional[float] = None,
        file_size: Optional[int] = None,
    ):
        self.entries: List[GlossaryEntry] = entries or []
        self.file_path: str = file_path
        self.file_sha256: Optional[str] = file_sha256
        self.file_mtime: Optional[float] = file_mtime
        self.file_size: Optional[int] = file_size
        self._key_to_entry: Dict[str, GlossaryEntry] = {}
        self._raw_terms_sorted_by_len: List[Tuple[str, GlossaryEntry]] = []
        self._build_index()

    def _build_index(self):
        """検索用インデックスの構築"""
        self._key_to_entry.clear()
        raw_list = []

        for entry in self.entries:
            # 代表語と類似語をすべて登録
            variants = entry.all_variants()
            for v in variants:
                if not v:
                    continue
                k = normalize_term_key(v)
                if k:
                    entry.normalized_keys.add(k)
                    self._key_to_entry[k] = entry
                raw_list.append((v, entry))

        # 最長一致検索用に文字列長の降順でソート
        self._raw_terms_sorted_by_len = sorted(raw_list, key=lambda x: len(x[0]), reverse=True)

    @classmethod
    def clear_cache(cls):
        """キャッシュの明示的クリア"""
        cls._cache.clear()

    @classmethod
    def from_file(cls, file_path: str, use_cache: bool = True) -> "GlossaryDictionary":
        """
        Excel (.xlsx) または CSV (.csv) ファイルから辞書を生成する。
        ハッシュおよびmtime/sizeによる差分検知を行い、変更がない場合はキャッシュを即座に返却する。
        """
        if not os.path.exists(file_path):
            return cls(entries=[], file_path=file_path)

        path_obj = Path(file_path).resolve()
        abs_path = str(path_obj)
        ext = path_obj.suffix.lower()

        try:
            stat = os.stat(abs_path)
            current_mtime = stat.st_mtime
            current_size = stat.st_size
        except OSError:
            return cls(entries=[], file_path=abs_path)

        if use_cache and abs_path in cls._cache:
            cached_info = cls._cache[abs_path]
            # 1. mtime と size が一致している場合は最速キャッシュヒット (ハッシュ計算すらスキップ)
            if cached_info.get("mtime") == current_mtime and cached_info.get("size") == current_size:
                return cached_info["instance"]

            # 2. mtime または size が変化している場合、SHA-256 ハッシュを計算して内容差分を確認
            current_hash = compute_file_sha256(abs_path)
            if cached_info.get("sha256") == current_hash:
                # 内容自体は変更されていない場合、メタデータを更新してキャッシュを返す
                cached_info["mtime"] = current_mtime
                cached_info["size"] = current_size
                return cached_info["instance"]
        else:
            current_hash = compute_file_sha256(abs_path)

        # 3. ファイルのパース実行
        if ext == ".xlsx":
            instance = cls._from_excel(abs_path)
        elif ext in [".csv", ".tsv", ".txt"]:
            instance = cls._from_csv(abs_path)
        else:
            instance = cls(entries=[], file_path=abs_path)

        instance.file_sha256 = current_hash
        instance.file_mtime = current_mtime
        instance.file_size = current_size

        # キャッシュの更新
        if use_cache:
            cls._cache[abs_path] = {
                "mtime": current_mtime,
                "size": current_size,
                "sha256": current_hash,
                "instance": instance,
            }

        return instance

    @classmethod
    def _from_excel(cls, file_path: str) -> "GlossaryDictionary":
        """Excelファイル (.xlsx) の読み込み"""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        if ws is None:
            return cls(entries=[], file_path=file_path)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return cls(entries=[], file_path=file_path)

        # ヘッダー行の解析
        header_row = rows[0]
        term_idx, syn_idx, desc_idx = cls._resolve_column_indices(header_row)

        entries: List[GlossaryEntry] = []
        for row in rows[1:]:
            if not row:
                continue

            term_raw = str(row[term_idx]).strip() if term_idx < len(row) and row[term_idx] is not None else ""
            if not term_raw or term_raw == "None":
                continue

            # 専門用語列内にカンマ区切りが含まれている場合は自動分割
            term_tokens = cls._parse_synonyms(term_raw)
            if not term_tokens:
                continue

            main_term = term_tokens[0]
            synonyms = term_tokens[1:]

            # 類似語列が別途ある場合はマージ
            if syn_idx != -1 and syn_idx < len(row) and row[syn_idx] is not None:
                syn_raw = str(row[syn_idx]).strip()
                if syn_raw and syn_raw != "None":
                    for s in cls._parse_synonyms(syn_raw):
                        if s != main_term and s not in synonyms:
                            synonyms.append(s)

            desc_val = ""
            if desc_idx != -1 and desc_idx < len(row) and row[desc_idx] is not None:
                d_str = str(row[desc_idx]).strip()
                if d_str != "None":
                    desc_val = d_str

            entries.append(GlossaryEntry(
                term=main_term,
                synonyms=synonyms,
                description=desc_val
            ))

        return cls(entries=entries, file_path=file_path)

    @classmethod
    def _from_csv(cls, file_path: str) -> "GlossaryDictionary":
        """CSVファイル (.csv) の読み込み"""
        entries: List[GlossaryEntry] = []
        with open(file_path, mode="r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = list(reader)
            if not rows:
                return cls(entries=[], file_path=file_path)

            header_row = rows[0]
            term_idx, syn_idx, desc_idx = cls._resolve_column_indices(header_row)

            for row in rows[1:]:
                if not row:
                    continue

                term_raw = row[term_idx].strip() if term_idx < len(row) else ""
                if not term_raw:
                    continue

                # 専門用語列内にカンマ区切りが含まれている場合は自動分割
                term_tokens = cls._parse_synonyms(term_raw)
                if not term_tokens:
                    continue

                main_term = term_tokens[0]
                synonyms = term_tokens[1:]

                # 類似語列が別途ある場合はマージ
                if syn_idx != -1 and syn_idx < len(row):
                    syn_raw = row[syn_idx].strip()
                    if syn_raw:
                        for s in cls._parse_synonyms(syn_raw):
                            if s != main_term and s not in synonyms:
                                synonyms.append(s)

                desc_val = ""
                if desc_idx != -1 and desc_idx < len(row):
                    desc_val = row[desc_idx].strip()

                entries.append(GlossaryEntry(
                    term=main_term,
                    synonyms=synonyms,
                    description=desc_val
                ))

        return cls(entries=entries, file_path=file_path)

    @staticmethod
    def _resolve_column_indices(header: List[Any]) -> Tuple[int, int, int]:
        """
        ヘッダー行から (Term, Synonyms, Description) の列インデックスを特定。
        2列構成の場合は Synonyms を -1 とし、Description を特定する。
        """
        term_idx = 0
        syn_idx = -1
        desc_idx = 1 if len(header) <= 2 else 2

        for i, col in enumerate(header):
            if col is None:
                continue
            col_str = str(col).strip().lower()
            if any(k in col_str for k in ["専門用語", "代表語", "term", "単語", "用語"]):
                term_idx = i
            elif any(k in col_str for k in ["類似語", "同義語", "略称", "別名", "synonym", "alias"]):
                syn_idx = i
            elif any(k in col_str for k in ["意味", "解説", "説明", "description", "備考", "詳細"]):
                desc_idx = i

        return term_idx, syn_idx, desc_idx

    @staticmethod
    def _parse_synonyms(syn_str: str) -> List[str]:
        """カンマや読点区切りの類似語文字列をリストに分割"""
        if not syn_str or syn_str == "None":
            return []
        tokens = re.split(r"[,、，\n\t]+", syn_str)
        res = []
        for t in tokens:
            cleaned = t.strip().strip("'\"")
            if cleaned and cleaned not in res:
                res.append(cleaned)
        return res

    @classmethod
    def save_to_excel(cls, file_path: str, entries: Union[List[GlossaryEntry], List[Dict[str, Any]]]) -> None:
        """
        専門用語辞書エントリを2列フォーマットのExcelファイル (.xlsx) として書き込み保存する。
        保存後に即座にハッシュキャッシュを更新する。
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Glossary"

        # ヘッダーの書き込み
        headers = ["専門用語", "意味・解説"]
        ws.append(headers)

        # スタイルの定義
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
        data_font = Font(name="Segoe UI", size=10, color="0F172A")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        alignment_center = Alignment(vertical="center")

        # ヘッダースタイルの適用
        for col_num in range(1, 3):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = alignment_center

        ws.row_dimensions[1].height = 26

        # データ行の書き込み
        for row_idx, item in enumerate(entries, start=2):
            if isinstance(item, GlossaryEntry):
                terms_text = ", ".join(item.all_variants())
                desc_text = item.description or ""
            elif isinstance(item, dict):
                terms_text = item.get("terms") or ""
                if not terms_text:
                    t = item.get("term", "")
                    syns = item.get("synonyms", [])
                    all_t = [t] + [s for s in syns if s and s != t]
                    terms_text = ", ".join(all_t)
                desc_text = item.get("description", "")
            else:
                continue

            ws.append([terms_text, desc_text])
            ws.row_dimensions[row_idx].height = 22

            for col_num in range(1, 3):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = alignment_center

        # 列幅の設定（視認性向上）
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 65

        # ディレクトリが存在しない場合は作成
        path_obj = Path(file_path).resolve()
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path_obj))

        # キャッシュの即時最新化
        abs_path = str(path_obj)
        if abs_path in cls._cache:
            del cls._cache[abs_path]
        # 最新の辞書を読み込み直してキャッシュ
        cls.from_file(abs_path, use_cache=True)

    def find_by_term(self, term: str) -> Optional[GlossaryEntry]:
        """代表語または類似語からエントリを検索"""
        k = normalize_term_key(term)
        return self._key_to_entry.get(k)

    def detect_terms(self, text: str) -> List[GlossaryEntry]:
        """
        自然文テキストから登録されている専門用語・類似語を最長一致で検知する。
        """
        if not text or not self.entries:
            return []

        detected_entries: List[GlossaryEntry] = []
        seen_entry_ids = set()

        # 1. 生の文字列による直接検索（最長一致）
        for raw_word, entry in self._raw_terms_sorted_by_len:
            if len(raw_word) < 2:
                continue
            # 単語が含まれているか（大文字小文字無視）
            pattern = re.escape(raw_word)
            if re.search(pattern, text, re.IGNORECASE):
                entry_id = id(entry)
                if entry_id not in seen_entry_ids:
                    seen_entry_ids.add(entry_id)
                    detected_entries.append(entry)

        # 2. 表記揺れ（ハイフン有無・全角半角）を検知するための正規化検索
        normalized_text = normalize_term_key(text)
        for k, entry in self._key_to_entry.items():
            if len(k) >= 2 and k in normalized_text:
                entry_id = id(entry)
                if entry_id not in seen_entry_ids:
                    seen_entry_ids.add(entry_id)
                    detected_entries.append(entry)

        return detected_entries

    def build_enriched_query(self, query: str) -> str:
        """
        自然言語クエリに対して、検知された専門用語の類似語・解説を自然な形で補強した文字列を生成する。
        """
        if not query or not self.entries:
            return query

        detected = self.detect_terms(query)
        if not detected:
            return query

        # 各用語の補足情報を構成
        supplements = []
        for entry in detected:
            parts = []
            # 代表語および同義語のうち、クエリに含まれていない関連語を追加
            for v in entry.all_variants():
                if v and v not in query:
                    parts.append(v)
            if entry.description:
                parts.append(entry.description)

            if parts:
                info_str = f"{entry.term} => {', '.join(parts[:4])}"
                supplements.append(info_str)

        if not supplements:
            return query

        # クエリ先頭に文脈コンテキスト [用語補足: ...] を付与してモデルに渡す
        context_header = f"[関連用語: {'; '.join(supplements)}]"
        return f"{context_header} {query}"

    def extract_enrichment_for_text(self, text: str) -> Tuple[List[str], List[str]]:
        """
        チャンク本文に含まれる用語から、メタデータヘッダーに注入すべき (aliases, context_notes) を抽出する。
        """
        if not text or not self.entries:
            return [], []

        detected = self.detect_terms(text)
        aliases: List[str] = []
        context_notes: List[str] = []

        for entry in detected:
            for v in entry.all_variants():
                if v and v not in aliases:
                    aliases.append(v)
            if entry.description and entry.description not in context_notes:
                context_notes.append(f"{entry.term}: {entry.description}")

        return aliases, context_notes
