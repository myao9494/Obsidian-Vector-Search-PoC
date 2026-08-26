"""
専門用語・類似語辞書モジュール (Glossary / Synonyms) 単体テスト
仕様:
- Excel (.xlsx) および CSV (.csv) ファイルから用語・類似語・説明を読み込めること。
- カンマ区切り（半角/全角）の類似語を正しくリスト化できること。
- 大文字/小文字、全角/半角、ハイフン有無（PJ-X ⇔ PJX ⇔ ｐｊｘ）の表記揺れを吸収できること。
- 自然言語文（質問文・文章）から登録用語を最長一致で正しく検知できること。
- クエリ補強テキスト（Embedding用）を生成できること。
"""

import os
import tempfile
import pytest
import openpyxl
from app.dictionary import GlossaryDictionary, GlossaryEntry


@pytest.fixture
def sample_excel_path():
    """テスト用Excel辞書ファイルの生成フィクスチャ"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["専門用語", "類似語", "意味・解説"])
    ws.append(["PJ-X", "プロジェクトX, PJX, PX", "2024年発足の社内基幹システム刷新プロジェクト"])
    ws.append(["ポチッと君", "ポチット、pochito", "社内の交通費・経費精算および旅費申請システム"])
    ws.append(["SLA", "サービスレベルアグリーメント, サービス品質保証", "契約上のシステム稼働率および品質保証基準"])
    ws.append(["DB", "データベース, Database", ""])  # 説明なしケース
    wb.save(path)

    yield path

    if os.path.exists(path):
        os.remove(path)


@pytest.fixture
def sample_csv_path():
    """テスト用CSV辞書ファイルの生成フィクスチャ"""
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w", encoding="utf-8") as f:
        f.write("Term,Synonyms,Description\n")
        f.write('PJ-X,"プロジェクトX, PJX",社内基幹刷新PJ\n')
        f.write("KVS,Redis,高速キーバリューストア\n")
        path = f.name

    yield path

    if os.path.exists(path):
        os.remove(path)


def test_load_excel_dictionary(sample_excel_path):
    """Excel辞書の読み込みテスト"""
    glossary = GlossaryDictionary.from_file(sample_excel_path)
    assert len(glossary.entries) == 4
    
    pjx_entry = glossary.find_by_term("PJ-X")
    assert pjx_entry is not None
    assert "プロジェクトX" in pjx_entry.synonyms
    assert "PJX" in pjx_entry.synonyms
    assert "PX" in pjx_entry.synonyms
    assert "社内基幹システム刷新" in pjx_entry.description


def test_load_csv_dictionary(sample_csv_path):
    """CSV辞書の読み込みテスト"""
    glossary = GlossaryDictionary.from_file(sample_csv_path)
    assert len(glossary.entries) == 2
    
    kvs_entry = glossary.find_by_term("KVS")
    assert kvs_entry is not None
    assert "Redis" in kvs_entry.synonyms
    assert "高速キーバリューストア" in kvs_entry.description


def test_detect_terms_in_natural_query(sample_excel_path):
    """自然言語の質問文からの用語検知テスト"""
    glossary = GlossaryDictionary.from_file(sample_excel_path)
    
    # 1. 略称 PJX を含む自然文
    query1 = "PJXのプロジェクトでの議事録ってどんなものがあったっけ"
    detected1 = glossary.detect_terms(query1)
    assert len(detected1) >= 1
    assert detected1[0].term == "PJ-X"
    
    # 2. 表記揺れ（小文字/ハイフンなし/全角）
    query2 = "ｐｊｘについての資料を探して"
    detected2 = glossary.detect_terms(query2)
    assert len(detected2) >= 1
    assert detected2[0].term == "PJ-X"

    # 3. 複数の専門用語を含む自然文
    query3 = "ポチッと君で交通費申請する際のSLA基準はどうなっていますか？"
    detected3 = glossary.detect_terms(query3)
    detected_terms = [e.term for e in detected3]
    assert "ポチッと君" in detected_terms
    assert "SLA" in detected_terms


def test_build_enriched_query(sample_excel_path):
    """Embedding用に文脈を補強したクエリ文字列の生成テスト"""
    glossary = GlossaryDictionary.from_file(sample_excel_path)
    
    query = "PJXのプロジェクトでの議事録ってどんなものがあったっけ"
    enriched = glossary.build_enriched_query(query)
    
    # 元のクエリ文の要素を保ちつつ、同義語・解説が自然に付加されていること
    assert "PJX" in enriched
    assert "プロジェクトX" in enriched
    assert "議事録" in enriched
    assert "どんなものがあったっけ" in enriched


def test_enrich_chunk_metadata(sample_excel_path):
    """チャンク本文に出現する用語からメタデータ（Aliases/Context）を補完するテスト"""
    glossary = GlossaryDictionary.from_file(sample_excel_path)
    
    chunk_text = "今日はプロジェクトXのキックオフを実施した。"
    aliases, context_notes = glossary.extract_enrichment_for_text(chunk_text)
    
    assert "PJ-X" in aliases or "PJX" in aliases or "PX" in aliases
    assert any("社内基幹システム刷新" in desc for desc in context_notes)


def test_load_two_column_excel_dictionary():
    """第1列にカンマ区切りの類似語を含む新2列フォーマットExcelの読み込みテスト"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glossary"
    # 2列フォーマット: [専門用語, 意味・解説]
    ws.append(["専門用語", "意味・解説"])
    ws.append(["PJ-X, プロジェクトX, PJX, PX", "2024年発足の社内基幹システム刷新プロジェクト"])
    ws.append(["ポチッと君、ポチット、pochito", "社内の経費精算システム"])
    ws.append(["単独用語", "類似語がない単独の専門用語の説明"])
    wb.save(path)

    try:
        glossary = GlossaryDictionary.from_file(path)
        assert len(glossary.entries) == 3

        # PJ-X の検証
        pjx = glossary.find_by_term("PJ-X")
        assert pjx is not None
        assert pjx.term == "PJ-X"
        assert "プロジェクトX" in pjx.synonyms
        assert "PJX" in pjx.synonyms
        assert "PX" in pjx.synonyms
        assert "社内基幹システム刷新" in pjx.description

        # 表記揺れ検知の確認 (PJX や プロジェクトX でも検知できること)
        detected = glossary.detect_terms("プロジェクトXの仕様書")
        assert len(detected) == 1
        assert detected[0].term == "PJ-X"

        # ポチッと君 の検証（読点 `、` 区切り）
        pochito = glossary.find_by_term("ポチッと君")
        assert pochito is not None
        assert "ポチット" in pochito.synonyms
        assert "pochito" in pochito.synonyms

        # 単独用語の検証
        single = glossary.find_by_term("単独用語")
        assert single is not None
        assert len(single.synonyms) == 0
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_load_two_column_csv_dictionary():
    """第1列にカンマ区切りを含む新2列フォーマットCSVの読み込みテスト"""
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w", encoding="utf-8") as f:
        f.write("専門用語,意味・解説\n")
        f.write('"PJ-X, プロジェクトX, PJX",社内基幹刷新PJ\n')
        f.write("KVS,高速キーバリューストア\n")
        path = f.name

    try:
        glossary = GlossaryDictionary.from_file(path)
        assert len(glossary.entries) == 2

        pjx = glossary.find_by_term("PJ-X")
        assert pjx is not None
        assert "プロジェクトX" in pjx.synonyms
        assert "PJX" in pjx.synonyms
        assert "社内基幹刷新PJ" in pjx.description
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_save_to_excel_and_reload():
    """辞書エントリをExcelファイルに保存し、再読み込みできることのテスト"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    entries = [
        GlossaryEntry(
            term="PJ-X",
            synonyms=["プロジェクトX", "PJX", "PX"],
            description="社内基幹刷新PJ",
        ),
        GlossaryEntry(
            term="ポチッと君",
            synonyms=["ポチット", "pochito"],
            description="経費精算システム",
        ),
    ]

    try:
        # Excelとして保存
        GlossaryDictionary.save_to_excel(path, entries)
        assert os.path.exists(path)

        # 再読み込みして検証
        reloaded = GlossaryDictionary.from_file(path)
        assert len(reloaded.entries) == 2

        pjx = reloaded.find_by_term("PJ-X")
        assert pjx is not None
        assert "プロジェクトX" in pjx.synonyms
        assert "PJX" in pjx.synonyms
        assert pjx.description == "社内基幹刷新PJ"

        pochito = reloaded.find_by_term("ポチッと君")
        assert pochito is not None
        assert "ポチット" in pochito.synonyms
        assert pochito.description == "経費精算システム"
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_dictionary_hash_cache_hit_and_invalidation():
    """ハッシュおよびmtime差分検知によるキャッシュヒットと、ファイル変更時の自動再パース検証"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["専門用語", "意味・解説"])
    ws.append(["TermA, 同義語A", "説明A"])
    wb.save(path)

    try:
        # 1回目: 初回読み込み & キャッシュ作成
        dict1 = GlossaryDictionary.from_file(path)
        assert len(dict1.entries) == 1
        assert dict1.find_by_term("TermA") is not None
        assert dict1.file_sha256 is not None

        # 2回目: ファイル未変更時 ➔ キャッシュが返却される（同一インスタンス）
        dict2 = GlossaryDictionary.from_file(path)
        assert dict2 is dict1  # キャッシュから即座に返却

        # 人間によるファイル更新シミュレーション（内容変更）
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Glossary"
        ws2.append(["専門用語", "意味・解説"])
        ws2.append(["TermA, 同義語A", "説明A"])
        ws2.append(["TermB, 同義語B", "説明B (新しく追記)"])
        wb2.save(path)

        # 3回目: ファイル変更後 ➔ ハッシュ差分を検知して自動で新辞書が再生成される
        dict3 = GlossaryDictionary.from_file(path)
        assert dict3 is not dict1
        assert len(dict3.entries) == 2
        assert dict3.find_by_term("TermB") is not None
        assert dict3.find_by_term("TermB").description == "説明B (新しく追記)"

        # 4回目: 再度未変更 ➔ 新しいキャッシュが返却される
        dict4 = GlossaryDictionary.from_file(path)
        assert dict4 is dict3
    finally:
        if os.path.exists(path):
            os.remove(path)


